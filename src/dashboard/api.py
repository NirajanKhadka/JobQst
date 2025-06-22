import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from fastapi import FastAPI, HTTPException, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from rich.console import Console

# Initialize console for rich output
console = Console()

# Initialize FastAPI app
app = FastAPI(title="AutoJobAgent Dashboard")

# Set up paths
BASE_DIR = Path(__file__).parent
TEMPLATES_DIR = BASE_DIR / "templates"
OUTPUT_DIR = BASE_DIR / "output"
LOGS_DIR = OUTPUT_DIR / "logs"
IPC_FILE = OUTPUT_DIR / "ipc.json"

# Ensure directories exist
OUTPUT_DIR.mkdir(exist_ok=True)
LOGS_DIR.mkdir(exist_ok=True)

# Set up templates
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# Mount static files (if they exist)
STATIC_DIR = BASE_DIR / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# Import real-time cache
try:
    from src.dashboard.job_cache import get_latest_jobs, get_cache_stats, job_cache
    CACHE_AVAILABLE = True
    console.print("[green]✅ Real-time cache integration enabled[/green]")
except ImportError:
    CACHE_AVAILABLE = False
    console.print("[yellow]⚠️ Real-time cache not available[/yellow]")

# Import job filtering system
try:
    from src.utils.job_filters import filter_job, filter_jobs_batch, get_filter_stats
    FILTER_AVAILABLE = True
    console.print("[green]✅ Job filtering integration enabled[/green]")
except ImportError:
    FILTER_AVAILABLE = False
    console.print("[yellow]⚠️ Job filtering not available[/yellow]")

# WebSocket connection manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: Dict):
        for connection in self.active_connections:
            await connection.send_json(message)


# Initialize connection manager
manager = ConnectionManager()


# Models
class PauseRequest(BaseModel):
    pause: bool = True


# Helper functions
def get_comprehensive_stats() -> Dict:
    """
    Get comprehensive job and application statistics.

    Returns:
        Dictionary with comprehensive statistics
    """
    try:
        from src.core.job_database import get_job_db
        from src.core import utils

        print("Getting comprehensive stats...")

        # Get available profiles
        profiles = utils.get_available_profiles()
        print(f"Available profiles for stats: {profiles}")

        comprehensive_stats = {
            "job_stats": {},
            "application_stats": {},
            "duplicate_stats": {},
            "performance_stats": {},
            "profile_stats": {},
            "recent_activity": [],
            "realtime_cache": {}  # New section for real-time cache stats
        }

        # Add real-time cache statistics if available
        if CACHE_AVAILABLE:
            try:
                cache_stats = get_cache_stats()
                comprehensive_stats["realtime_cache"] = {
                    "enabled": True,
                    "cache_size": cache_stats.get("cache_size", 0),
                    "total_added": cache_stats.get("total_added", 0),
                    "total_retrieved": cache_stats.get("total_retrieved", 0),
                    "cache_hits": cache_stats.get("cache_hits", 0),
                    "last_updated": cache_stats.get("last_updated"),
                    "max_size": cache_stats.get("max_size", 500)
                }
                console.print(f"[blue]🚀 Real-time cache stats: {cache_stats.get('cache_size', 0)} jobs cached[/blue]")
            except Exception as e:
                console.print(f"[yellow]⚠️ Error getting cache stats: {e}[/yellow]")
                comprehensive_stats["realtime_cache"] = {"enabled": False, "error": str(e)}
        else:
            comprehensive_stats["realtime_cache"] = {"enabled": False}

        # First, get stats from the default database
        try:
            print("Getting stats for default database")
            default_db = get_job_db()  # Default database
            default_stats = default_db.get_stats()
            print(f"Default database stats: {default_stats}")
            
            # Add default database stats to totals
            total_jobs = default_stats.get("total_jobs", 0)
            total_applied = default_stats.get("applied_jobs", 0)
            
            # Store default database stats
            comprehensive_stats["profile_stats"]["default"] = {
                "jobs": default_stats,
                "duplicates": {"duplicate_jobs": 0}
            }
            
        except Exception as e:
            print(f"Warning: Could not get stats for default database: {e}")

        # Aggregate stats across all profiles
        total_jobs = 0
        total_applied = 0
        total_duplicates = 0
        total_unique_companies = 0

        for profile_name in profiles:
            try:
                print(f"Getting stats for profile: {profile_name}")

                # Get job database stats
                db = get_job_db(profile_name)
                job_stats = db.get_stats()
                print(f"Job stats for {profile_name}: {job_stats}")

                # Get duplicate information
                duplicates = db.get_duplicates()
                duplicate_count = len(duplicates)
                print(f"Duplicate count for {profile_name}: {duplicate_count}")

                # Store per-profile stats
                comprehensive_stats["profile_stats"][profile_name] = {
                    "jobs": job_stats,
                    "duplicates": {"duplicate_jobs": duplicate_count}
                }

                # Aggregate totals
                total_jobs += job_stats.get("total_jobs", 0)
                total_applied += job_stats.get("applied_jobs", 0)
                total_duplicates += duplicate_count
                total_unique_companies += job_stats.get("unique_companies", 0)

            except Exception as e:
                print(f"Warning: Could not get stats for profile {profile_name}: {e}")

        # Overall job statistics
        comprehensive_stats["job_stats"] = {
            "total_jobs": total_jobs,  # Fixed field name
            "total_scraped": total_jobs,
            "unique_jobs": total_jobs - total_duplicates,
            "duplicates_detected": total_duplicates,
            "applied_jobs": total_applied,
            "unapplied_jobs": total_jobs - total_applied,
            "unique_companies": total_unique_companies,
            "success_rate": round((total_applied / total_jobs * 100), 2) if total_jobs > 0 else 0
        }

        # Get application statistics from Excel logs
        app_stats = get_application_stats()
        comprehensive_stats["application_stats"] = app_stats

        # Duplicate detection statistics
        comprehensive_stats["duplicate_stats"] = {
            "total_duplicates": total_duplicates,
            "duplicate_rate": round((total_duplicates / total_jobs * 100), 2) if total_jobs > 0 else 0,
            "unique_rate": round(((total_jobs - total_duplicates) / total_jobs * 100), 2) if total_jobs > 0 else 0
        }

        # Performance statistics
        comprehensive_stats["performance_stats"] = {
            "profiles_active": len(profiles),
            "avg_jobs_per_profile": round(total_jobs / len(profiles), 2) if profiles else 0,
            "system_health": "Excellent" if total_jobs > 0 else "No Data"
        }

        return comprehensive_stats

    except Exception as e:
        console.print(f"[red]Error getting comprehensive stats: {e}[/red]")
        return {
            "job_stats": {"total_scraped": 0, "unique_jobs": 0, "duplicates_detected": 0, "applied_jobs": 0, "unapplied_jobs": 0, "success_rate": 0},
            "application_stats": {"Applied": 0, "Skipped": 0, "Failed": 0, "Manual": 0, "Total": 0},
            "duplicate_stats": {"total_duplicates": 0, "duplicate_rate": 0, "unique_rate": 100},
            "performance_stats": {"profiles_active": 0, "avg_jobs_per_profile": 0, "system_health": "Error"},
            "profile_stats": {},
            "recent_activity": [],
            "realtime_cache": {"enabled": False, "error": str(e)}
        }


def get_application_stats() -> Dict:
    """
    Get application statistics from the log file.

    Returns:
        Dictionary with application statistics
    """
    stats = {
        "Applied": 0,
        "Skipped": 0,
        "Failed": 0,
        "Manual": 0,
        "Total": 0,
        "manual_reasons": {},
        "failure_reasons": {},
        "recent_applications": []
    }

    log_file = LOGS_DIR / "applications.xlsx"

    if not log_file.exists():
        return stats

    try:
        wb = openpyxl.load_workbook(log_file, read_only=True)
        ws = wb.active

        # Get headers
        headers = [cell.value for cell in ws[1]]

        # Skip header row
        rows = list(ws.rows)[1:]

        # Count statuses and analyze reasons
        for row in rows:
            if len(row) > 6:
                status = row[6].value  # Status is in column G (index 6)
                if status:
                    status = str(status)
                    stats["Total"] += 1

                    # Create row dict for recent applications
                    row_data = {}
                    for i, cell in enumerate(row):
                        if i < len(headers) and headers[i]:
                            if isinstance(cell.value, datetime):
                                row_data[headers[i]] = cell.value.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                row_data[headers[i]] = cell.value

                    if "Applied" in status:
                        stats["Applied"] += 1
                    elif "Skipped" in status:
                        stats["Skipped"] += 1
                    elif "Failed" in status:
                        stats["Failed"] += 1
                        # Track failure reasons
                        reason = row_data.get("Notes", "Unknown")
                        if reason:
                            stats["failure_reasons"][reason] = stats["failure_reasons"].get(reason, 0) + 1
                    elif "Manual" in status:
                        stats["Manual"] += 1
                        # Track manual reasons
                        reason = row_data.get("Notes", "Manual review required")
                        if reason:
                            stats["manual_reasons"][reason] = stats["manual_reasons"].get(reason, 0) + 1

                    # Add to recent applications (last 10)
                    if len(stats["recent_applications"]) < 10:
                        stats["recent_applications"].append(row_data)

        return stats

    except Exception as e:
        console.print(f"[bold red]Error getting application stats: {e}[/bold red]")
        return stats


def get_recent_logs(limit: int = 100) -> List[Dict]:
    """
    Get recent application logs.
    
    Args:
        limit: Maximum number of logs to return
        
    Returns:
        List of log entries
    """
    logs = []
    log_file = LOGS_DIR / "applications.xlsx"
    
    if not log_file.exists():
        return logs
    
    try:
        wb = openpyxl.load_workbook(log_file, read_only=True)
        ws = wb.active
        
        # Get all rows and reverse to get most recent first
        rows = list(ws.rows)
        headers = [cell.value for cell in rows[0]]
        
        # Get data rows (skip header), most recent first
        data_rows = rows[1:][::-1][:limit]
        
        # Convert rows to dictionaries
        for row in data_rows:
            log_entry = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    # Format timestamp if it's a datetime
                    if i == 0 and isinstance(cell.value, datetime):
                        log_entry[headers[i]] = cell.value.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        log_entry[headers[i]] = cell.value
            logs.append(log_entry)
        
        return logs
    
    except Exception as e:
        console.print(f"[bold red]Error getting recent logs: {e}[/bold red]")
        return logs


def get_pause_status() -> bool:
    """
    Get the current pause status.
    
    Returns:
        True if paused, False otherwise
    """
    if not IPC_FILE.exists():
        return False
    
    try:
        with open(IPC_FILE, "r") as f:
            ipc_data = json.load(f)
        
        return ipc_data.get("pause", False)
    
    except (json.JSONDecodeError, FileNotFoundError):
        return False


def set_pause_status(pause: bool) -> bool:
    """
    Set the pause status.
    
    Args:
        pause: True to pause, False to resume
        
    Returns:
        True if successful, False otherwise
    """
    try:
        with open(IPC_FILE, "w") as f:
            json.dump({"pause": pause}, f)
        
        return True
    
    except Exception as e:
        console.print(f"[bold red]Error setting pause status: {e}[/bold red]")
        return False


# Routes
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """
    Render the enhanced dashboard page.
    """
    return templates.TemplateResponse(
        "dashboard.html",
        {"request": request, "title": "AutoJobAgent Dashboard"}
    )


@app.get("/api-test", response_class=HTMLResponse)
async def api_test_page(request: Request):
    """
    Render the API test page for testing dashboard endpoints.
    """
    return templates.TemplateResponse(
        "api_test.html",
        {"request": request, "title": "Dashboard API Test Center"}
    )


@app.get("/stats")
async def stats():
    """
    Get application statistics (legacy endpoint).
    """
    return get_application_stats()


@app.get("/api/comprehensive-stats")
async def comprehensive_stats():
    """
    Get comprehensive job and application statistics.
    """
    return get_comprehensive_stats()


@app.get("/api/job-stats")
async def job_stats():
    """
    Get job scraping and database statistics.
    """
    stats = get_comprehensive_stats()
    return stats.get("job_stats", {})


@app.get("/api/application-stats")
async def application_stats():
    """
    Get detailed application statistics with reasons.
    """
    stats = get_comprehensive_stats()
    return stats.get("application_stats", {})


@app.get("/api/duplicate-stats")
async def duplicate_stats():
    """
    Get duplicate detection statistics.
    """
    stats = get_comprehensive_stats()
    return stats.get("duplicate_stats", {})


@app.get("/api/performance-stats")
async def performance_stats():
    """
    Get system performance statistics.
    """
    stats = get_comprehensive_stats()
    return stats.get("performance_stats", {})


@app.get("/api/profile-stats")
async def profile_stats():
    """
    Get per-profile statistics.
    """
    stats = get_comprehensive_stats()
    return stats.get("profile_stats", {})


@app.get("/api/health")
async def health_check():
    """Simple health check endpoint."""
    return {"status": "healthy", "message": "Dashboard API is running"}

@app.get("/api/debug-dashboard")
async def debug_dashboard():
    """Debug endpoint to test dashboard components."""
    try:
        from src.core.job_database import get_job_db
        import utils

        # Test basic components
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles found", "profiles": []}

        profile_name = profiles[0]
        db = get_job_db(profile_name)
        stats = db.get_stats()

        return {
            "status": "success",
            "profiles": profiles,
            "profile_used": profile_name,
            "stats": stats,
            "message": "Basic components working"
        }
    except Exception as e:
        import traceback
        return {
            "status": "error",
            "error": str(e),
            "traceback": traceback.format_exc()
        }

@app.get("/api/simple-dashboard-numbers")
async def simple_dashboard_numbers():
    """Simplified dashboard numbers endpoint for testing."""
    try:
        # Get basic stats only
        stats = get_comprehensive_stats()
        job_stats = stats.get("job_stats", {})

        return {
            "timestamp": "2025-06-20T00:00:00",
            "total_jobs": job_stats.get("total_jobs", 0),
            "applied_jobs": job_stats.get("applied_jobs", 0),
            "pending_jobs": job_stats.get("unapplied_jobs", 0),
            "unique_companies": job_stats.get("unique_companies", 0),
            "success_rate": job_stats.get("success_rate", 0),
            "status": "working"
        }
    except Exception as e:
        import traceback
        return {
            "error": str(e),
            "traceback": traceback.format_exc(),
            "status": "error"
        }

@app.get("/api/dashboard-numbers")
async def dashboard_numbers():
    """
    Get real-time dashboard numbers and statistics for quick overview.
    This endpoint now uses the centralized `get_comprehensive_stats` function.
    """
    try:
        console.print("[cyan]📊 Getting comprehensive dashboard numbers...[/cyan]")

        # Get comprehensive stats which now includes all databases
        stats = get_comprehensive_stats()

        console.print(f"[green]✅ Comprehensive stats retrieved: {stats['job_stats']['total_jobs']} jobs[/green]")

        return stats

    except Exception as e:
        console.print(f"[red]❌ Error getting dashboard numbers: {e}[/red]")
        import traceback
        return JSONResponse(
            status_code=500,
            content={
                "error": str(e),
                "traceback": traceback.format_exc(),
            },
        )


@app.get("/api/jobs")
async def get_jobs(
    limit: int = 50,
    offset: int = 0,
    profile: str = None,
    site: str = None,
    search: str = None,
    applied: bool = None
):
    """
    Get scraped jobs with filtering and pagination.

    Args:
        limit: Maximum number of jobs to return (default: 50, max: 200)
        offset: Number of jobs to skip for pagination
        profile: Filter by profile name
        site: Filter by job site (eluta, indeed, etc.)
        search: Search in job title or company
        applied: Filter by application status (true/false)
    """
    return await get_jobs_enhanced(limit, offset, profile, site, search, applied)


@app.get("/api/jobs-table")
async def get_jobs_table(
    limit: int = 100,
    offset: int = 0,
    profile: str = None,
    site: str = None,
    search: str = None,
    applied: bool = None
):
    """
    Get scraped jobs formatted for detailed table display.

    Args:
        limit: Maximum number of jobs to return (default: 100, max: 500)
        offset: Number of jobs to skip for pagination
        profile: Filter by profile name
        site: Filter by job site (eluta, indeed, etc.)
        search: Search in job title or company
        applied: Filter by application status (true/false)
    """
    return await get_jobs_enhanced(limit, offset, profile, site, search, applied, table_format=True)


async def get_jobs_enhanced(
    limit: int = 50,
    offset: int = 0,
    profile: str = None,
    site: str = None,
    search: str = None,
    applied: bool = None,
    experience: str = None,
    table_format: bool = False
):
    """
    Enhanced endpoint to get jobs with filtering, searching, and pagination.
    This version is more robust and handles missing data gracefully.
    """
    print(f"Dashboard API: get_jobs_enhanced called with limit={limit}, offset={offset}, profile={profile}, table_format={table_format}")

    try:
        from src.core.job_database import get_job_db
        import utils

        profiles = utils.get_available_profiles()
        print(f"Available profiles: {profiles}")
        if not profiles:
            return {"jobs": [], "total": 0}

        profile_name = profile if profile and profile in profiles else profiles[0]
        print(f"Using profile: {profile_name}")

        db = get_job_db(profile_name)
        print(f"Database connected: {db.db_path}")

        db_stats = db.get_stats()
        print(f"Database stats: {db_stats}")

        filters = {'site': site, 'applied': applied, 'experience': experience}
        print(f"Filters: {filters}")

        # This call should now work correctly with the updated get_jobs method
        jobs = db.get_jobs(limit=limit, offset=offset, filters=filters, search_query=search)
        
        enhanced_jobs = []
        for job in jobs:
            enhanced_job = dict(job)

            # Safely parse JSON data fields
            try:
                raw_data = json.loads(enhanced_job.get('raw_data', '{}') or '{}')
                analysis_data = json.loads(enhanced_job.get('analysis_data', '{}') or '{}')
            except (json.JSONDecodeError, TypeError):
                raw_data = {}
                analysis_data = {}

            # Use safe .get() with default values to prevent AttributeErrors
            requirements = analysis_data.get('requirements', {}) if analysis_data else {}
            experience_level = requirements.get('experience_level', 'Not specified') if requirements else 'Not specified'
            
            match_details = analysis_data.get('match_details', {}) if analysis_data else {}
            match_score = match_details.get('score', 0) if match_details else 0

            enhanced_job['experience'] = experience_level
            enhanced_job['match_score'] = match_score
            enhanced_job['applied_status'] = enhanced_job.get('applied', False) # Assuming 'applied' column
            
            enhanced_jobs.append(enhanced_job)
        
        print(f"Retrieved {len(enhanced_jobs)} jobs from database")

        total_count = db.get_job_count(filters=filters, search_query=search)

        return {
            "jobs": enhanced_jobs,
            "total": total_count,
            "has_more": len(jobs) > 0,
            "profile": profile_name,
            "filters": {
                "site": site,
                "search": search,
                "applied": applied,
                "experience": experience
            }
        }
    except Exception as e:
        print(f"Error getting jobs: {e}")
        import traceback
        traceback.print_exc()
        return {"jobs": [], "total": 0, "error": str(e)}


@app.get("/api/job/{job_id}")
async def get_job_details(job_id: int, profile: str = None):
    """
    Get detailed information for a specific job.

    Args:
        job_id: Job ID
        profile: Profile name (optional)
    """
    try:
        from src.core.job_database import get_job_db
        import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get job database
        db = get_job_db(profile_name)

        # Get job details
        job = db.get_job_by_id(job_id)

        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        # Enhance job details
        enhanced_job = dict(job)

        # Fix company name
        company = enhanced_job.get('company', 'Unknown Company')
        if company == 'Unknown Company' or not company:
            url = enhanced_job.get('url', '')
            if 'eluta.ca' in url and '/company/' in url:
                try:
                    company_part = url.split('/company/')[1].split('/')[0]
                    company = company_part.replace('-', ' ').title()
                except:
                    pass
        enhanced_job['company'] = company

        return enhanced_job

    except Exception as e:
        print(f"Error getting job details: {e}")
        return {"error": str(e)}


@app.delete("/api/jobs/clear")
async def clear_all_jobs(profile: str = None):
    """
    Clear all jobs from the database for a fresh start.

    Args:
        profile: Profile name (optional)
    """
    try:
        from src.core.job_database import get_job_db
        import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get job database
        db = get_job_db(profile_name)

        # Get current stats before clearing
        before_stats = db.get_stats()

        # Clear all jobs
        success = db.clear_all_jobs()

        if success:
            # Get stats after clearing
            after_stats = db.get_stats()

            # Broadcast update via WebSocket
            await manager.broadcast({
                "type": "jobs_cleared",
                "profile": profile_name,
                "jobs_removed": before_stats.get("total_jobs", 0),
                "remaining_jobs": after_stats.get("total_jobs", 0)
            })

            return {
                "success": True,
                "message": f"Cleared {before_stats.get('total_jobs', 0)} jobs from database",
                "jobs_removed": before_stats.get("total_jobs", 0),
                "profile": profile_name
            }
        else:
            return {"error": "Failed to clear jobs from database"}

    except Exception as e:
        console.print(f"[red]Error clearing jobs: {e}[/red]")
        return {"error": str(e)}


@app.post("/api/job/{job_id}/apply")
async def apply_to_job(job_id: int, profile: str = None):
    """
    Apply to a specific job.

    Args:
        job_id: Job ID
        profile: Profile name (optional)
    """
    try:
        from src.core.job_database import get_job_db
        import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get job database
        db = get_job_db(profile_name)

        # Get job details
        job = db.get_job_by_id(job_id)

        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        # Mark job as applied (for now, just update the database)
        # In a full implementation, this would trigger the actual application process
        db.mark_applied(job['url'], 'Applied via Dashboard')

        # Broadcast update via WebSocket
        await manager.broadcast({
            "type": "job_applied",
            "job_id": job_id,
            "job_title": job.get('title', 'Unknown'),
            "company": job.get('company', 'Unknown')
        })

        return {"success": True, "message": f"Application initiated for {job.get('title', 'job')}"}

    except Exception as e:
        console.print(f"[red]Error applying to job: {e}[/red]")
        return {"error": str(e)}


@app.delete("/api/job/{job_id}")
async def delete_job(job_id: int, profile: str = None):
    """
    Delete a specific job.

    Args:
        job_id: Job ID
        profile: Profile name (optional)
    """
    try:
        from src.core.job_database import get_job_db
        import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get job database
        db = get_job_db(profile_name)

        # Get job details before deletion
        job = db.get_job_by_id(job_id)

        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        # Delete the job
        success = db.delete_job(job_id)

        if success:
            # Broadcast update via WebSocket
            await manager.broadcast({
                "type": "job_deleted",
                "job_id": job_id,
                "job_title": job.get('title', 'Unknown'),
                "company": job.get('company', 'Unknown')
            })

            return {"success": True, "message": f"Job '{job.get('title', 'Unknown')}' deleted successfully"}
        else:
            return {"error": "Failed to delete job from database"}

    except Exception as e:
        console.print(f"[red]Error deleting job: {e}[/red]")
        return {"error": str(e)}


@app.get("/api/sites")
async def get_available_sites():
    """
    Get list of available job sites.
    """
    try:
        from src.core.job_database import get_job_db
        from src.core import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        
        # Use the first available profile or default
        profile_name = profiles[0] if profiles else None
        
        # Get database
        db = get_job_db(profile_name)
        
        # Get unique sites
        sites = db.get_unique_sites()
        
        return {
            "sites": sites,
            "total_sites": len(sites),
            "profile_used": profile_name
        }
        
    except Exception as e:
        console.print(f"[red]Error getting sites: {e}[/red]")
        return {"sites": [], "total_sites": 0, "error": str(e)}


@app.get("/api/system-status")
async def get_system_status():
    """
    Get real-time system status and health metrics.

    Returns:
        System health, database status, recent errors, performance metrics
    """
    try:
        from src.core.job_database import get_job_db
        import utils
        from datetime import datetime

        console.print("[cyan]🔍 Checking system status...[/cyan]")

        # Get available profiles
        profiles = utils.get_available_profiles()

        # Database connectivity check
        db_status = {}
        for profile_name in profiles:
            try:
                db = get_job_db(profile_name)
                stats = db.get_stats()
                db_status[profile_name] = {
                    "connected": True,
                    "total_jobs": stats.get("total_jobs", 0),
                    "last_updated": datetime.now().isoformat()
                }
            except Exception as e:
                db_status[profile_name] = {
                    "connected": False,
                    "error": str(e),
                    "last_updated": datetime.now().isoformat()
                }

        # System resource usage
        try:
            import psutil
            memory = psutil.virtual_memory()
            disk = psutil.disk_usage('.')
            cpu_percent = psutil.cpu_percent(interval=1)

            system_resources = {
                "memory": {
                    "total_gb": round(memory.total / (1024**3), 2),
                    "used_gb": round(memory.used / (1024**3), 2),
                    "available_gb": round(memory.available / (1024**3), 2),
                    "percent_used": memory.percent
                },
                "disk": {
                    "total_gb": round(disk.total / (1024**3), 2),
                    "used_gb": round(disk.used / (1024**3), 2),
                    "free_gb": round(disk.free / (1024**3), 2),
                    "percent_used": round((disk.used / disk.total) * 100, 2)
                },
                "cpu": {
                    "percent_used": cpu_percent,
                    "core_count": psutil.cpu_count()
                }
            }
        except (ImportError, Exception):
            # Fallback system info without psutil
            import shutil
            disk_usage = shutil.disk_usage('.')
            system_resources = {
                "memory": {"note": "psutil not available - install with: pip install psutil"},
                "disk": {
                    "total_gb": round(disk_usage.total / (1024**3), 2),
                    "used_gb": round((disk_usage.total - disk_usage.free) / (1024**3), 2),
                    "free_gb": round(disk_usage.free / (1024**3), 2),
                    "percent_used": round(((disk_usage.total - disk_usage.free) / disk_usage.total) * 100, 2)
                },
                "cpu": {"note": "psutil not available - install with: pip install psutil"}
            }

        # Check log files
        log_status = {}
        log_files = ["applications.xlsx", "fallback.csv"]
        for log_file in log_files:
            log_path = LOGS_DIR / log_file
            if log_path.exists():
                stat = log_path.stat()
                log_status[log_file] = {
                    "exists": True,
                    "size_mb": round(stat.st_size / (1024*1024), 2),
                    "modified": datetime.fromtimestamp(stat.st_mtime).isoformat()
                }
            else:
                log_status[log_file] = {"exists": False}

        # Overall health assessment
        health_issues = []
        if not profiles:
            health_issues.append("No profiles available")

        db_connected = sum(1 for status in db_status.values() if status.get("connected", False))
        if db_connected == 0:
            health_issues.append("No database connections")
        elif db_connected < len(profiles):
            health_issues.append(f"Only {db_connected}/{len(profiles)} databases connected")

        # Check if we have recent data
        total_jobs = sum(status.get("total_jobs", 0) for status in db_status.values() if status.get("connected", False))
        if total_jobs == 0:
            health_issues.append("No jobs in database")

        overall_health = "Healthy" if not health_issues else "Warning" if len(health_issues) <= 2 else "Critical"
        health_color = "green" if overall_health == "Healthy" else "yellow" if overall_health == "Warning" else "red"

        return {
            "timestamp": datetime.now().isoformat(),
            "overall_health": {
                "status": overall_health,
                "color": health_color,
                "issues": health_issues,
                "score": max(0, 100 - (len(health_issues) * 25))
            },
            "profiles": {
                "total": len(profiles),
                "available": profiles,
                "active": db_connected
            },
            "databases": db_status,
            "system_resources": system_resources,
            "log_files": log_status,
            "services": {
                "dashboard": {"status": "running", "port": 8001},
                "ollama": {"status": "checking", "note": "Check via main application"}
            }
        }

    except Exception as e:
        console.print(f"[red]❌ Error getting system status: {e}[/red]")
        return {
            "timestamp": datetime.now().isoformat(),
            "overall_health": {"status": "Error", "color": "red", "issues": [str(e)], "score": 0},
            "error": str(e)
        }


@app.get("/api/quick-test")
async def quick_test():
    """
    Quick test endpoint to verify dashboard API is working.

    Returns basic system info and connectivity status.
    """
    try:
        from datetime import datetime
        import utils

        # Test basic functionality
        profiles = utils.get_available_profiles()

        # Test database connectivity
        db_test = "Failed"
        if profiles:
            try:
                from src.core.job_database import get_job_db
                db = get_job_db(profiles[0])
                stats = db.get_stats()
                db_test = f"Connected - {stats.get('total_jobs', 0)} jobs"
            except Exception as e:
                db_test = f"Error: {str(e)[:50]}"

        # Test log file access
        log_test = "No logs"
        log_file = LOGS_DIR / "applications.xlsx"
        if log_file.exists():
            log_test = f"Found - {round(log_file.stat().st_size / 1024, 2)} KB"

        return {
            "status": "OK",
            "timestamp": datetime.now().isoformat(),
            "dashboard_api": "Working",
            "profiles_found": len(profiles),
            "profiles": profiles,
            "database_test": db_test,
            "log_file_test": log_test,
            "endpoints_available": [
                "/api/dashboard-numbers",
                "/api/comprehensive-stats",
                "/api/jobs-table",
                "/api/system-status",
                "/api/quick-test"
            ]
        }

    except Exception as e:
        return {
            "status": "ERROR",
            "timestamp": datetime.now().isoformat(),
            "error": str(e),
            "dashboard_api": "Failed"
        }


@app.get("/api/test-data")
@app.post("/api/test-data")
async def add_test_data():
    """
    Add sample test data for dashboard demonstration.
    """
    try:
        from src.core.job_database import get_job_db
        import utils
        from datetime import datetime, timedelta

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use first available profile
        profile_name = profiles[0]
        db = get_job_db(profile_name)

        # Sample job data with proper company names
        sample_jobs = [
            {
                "title": "Junior Data Analyst",
                "company": "Microsoft Canada",
                "location": "Toronto, ON",
                "url": "https://www.eluta.ca/job/microsoft-junior-data-analyst-toronto",
                "apply_url": "https://www.eluta.ca/job/microsoft-junior-data-analyst-toronto",
                "summary": "Join Microsoft's data team as a Junior Data Analyst. Work with Python, SQL, and Power BI to analyze business data and create insights.",
                "posted_date": "2 days ago",
                "salary": "$65,000 - $75,000",
                "job_type": "Full-time",
                "site": "eluta_enhanced",
                "search_keyword": "Data Analyst",
                "experience_level": "entry",
                "applied": False,
                "scraped_at": datetime.now().isoformat()
            },
            {
                "title": "Python Developer - Entry Level",
                "company": "Shopify Inc",
                "location": "Ottawa, ON",
                "url": "https://www.eluta.ca/job/shopify-python-developer-ottawa",
                "apply_url": "https://www.eluta.ca/job/shopify-python-developer-ottawa",
                "summary": "Entry-level Python Developer position at Shopify. Work on e-commerce solutions using Python, Django, and modern web technologies.",
                "posted_date": "1 day ago",
                "salary": "$70,000 - $80,000",
                "job_type": "Full-time",
                "site": "eluta_enhanced",
                "search_keyword": "Python Developer",
                "experience_level": "entry",
                "applied": True,
                "scraped_at": (datetime.now() - timedelta(days=1)).isoformat()
            },
            {
                "title": "Business Intelligence Analyst",
                "company": "Royal Bank of Canada",
                "location": "Toronto, ON",
                "url": "https://www.eluta.ca/job/rbc-business-intelligence-analyst-toronto",
                "apply_url": "https://www.eluta.ca/job/rbc-business-intelligence-analyst-toronto",
                "summary": "Join RBC's BI team to create dashboards and reports using Tableau, Power BI, and SQL. Great opportunity for recent graduates.",
                "posted_date": "3 days ago",
                "salary": "$60,000 - $70,000",
                "job_type": "Full-time",
                "site": "eluta_enhanced",
                "search_keyword": "Business Analyst",
                "experience_level": "entry",
                "applied": False,
                "scraped_at": (datetime.now() - timedelta(days=2)).isoformat()
            },
            {
                "title": "SQL Developer - Junior",
                "company": "Manulife Financial",
                "location": "Toronto, ON",
                "url": "https://www.eluta.ca/job/manulife-sql-developer-toronto",
                "apply_url": "https://www.eluta.ca/job/manulife-sql-developer-toronto",
                "summary": "Junior SQL Developer role at Manulife. Work with large datasets, create stored procedures, and optimize database performance.",
                "posted_date": "4 days ago",
                "salary": "$55,000 - $65,000",
                "job_type": "Full-time",
                "site": "eluta_enhanced",
                "search_keyword": "SQL Developer",
                "experience_level": "entry",
                "applied": False,
                "scraped_at": (datetime.now() - timedelta(days=3)).isoformat()
            },
            {
                "title": "Data Visualization Specialist",
                "company": "Deloitte Canada",
                "location": "Toronto, ON",
                "url": "https://www.eluta.ca/job/deloitte-data-visualization-specialist-toronto",
                "apply_url": "https://www.eluta.ca/job/deloitte-data-visualization-specialist-toronto",
                "summary": "Create compelling data visualizations using Tableau, Power BI, and D3.js. Work with clients to transform data into actionable insights.",
                "posted_date": "5 days ago",
                "salary": "$65,000 - $75,000",
                "job_type": "Full-time",
                "site": "eluta_enhanced",
                "search_keyword": "Data Visualization",
                "experience_level": "entry",
                "applied": True,
                "scraped_at": (datetime.now() - timedelta(days=4)).isoformat()
            }
        ]

        added_count = 0
        for job in sample_jobs:
            try:
                db.add_job(job)
                added_count += 1
            except Exception as e:
                console.print(f"[yellow]Warning: Could not add job {job['title']}: {e}[/yellow]")

        # Get updated stats
        stats = db.get_stats()

        # Broadcast update via WebSocket
        await manager.broadcast({
            "type": "test_data_added",
            "jobs_added": added_count,
            "total_jobs": stats["total_jobs"]
        })

        return {
            "success": True,
            "message": f"Added {added_count} sample jobs",
            "stats": stats
        }

    except Exception as e:
        console.print(f"[red]Error adding test data: {e}[/red]")
        return {"error": str(e)}


@app.get("/api/manual-reviews")
async def get_manual_reviews(status: str = "pending", profile: str = None):
    """
    Get manual review queue items.

    Args:
        status: Status filter (pending, resolved, all)
        profile: Profile name (optional)
    """
    try:
        from src.core.job_database import get_job_db
        import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get job database
        db = get_job_db(profile_name)

        # Get manual review items
        reviews = db.get_manual_review_queue(status)

        return {
            "reviews": reviews,
            "total": len(reviews),
            "status_filter": status,
            "profile": profile_name
        }

    except Exception as e:
        console.print(f"[red]Error getting manual reviews: {e}[/red]")
        return {"error": str(e)}


@app.get("/api/manual-review/{review_id}")
async def get_manual_review_details(review_id: int, profile: str = None):
    """
    Get detailed information for a specific manual review item.

    Args:
        review_id: Manual review ID
        profile: Profile name (optional)
    """
    try:
        from manual_review_manager import ManualReviewManager
        import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get manual review manager
        manager = ManualReviewManager(profile_name)

        # Get review details
        details = manager.get_review_details(review_id)

        if not details:
            raise HTTPException(status_code=404, detail="Manual review item not found")

        return details

    except Exception as e:
        console.print(f"[red]Error getting manual review details: {e}[/red]")
        return {"error": str(e)}


@app.get("/api/application/{application_id}")
async def get_application_details(application_id: int, profile: str = None):
    """
    Get detailed information for a specific application.

    Args:
        application_id: Application ID
        profile: Profile name (optional)
    """
    try:
        from src.core.job_database import get_job_db
        import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get job database
        db = get_job_db(profile_name)

        # Get application details
        details = db.get_application_details(application_id)

        if not details:
            raise HTTPException(status_code=404, detail="Application not found")

        return details

    except Exception as e:
        console.print(f"[red]Error getting application details: {e}[/red]")
        return {"error": str(e)}


@app.get("/api/failed-applications")
async def get_failed_applications(profile: str = None):
    """
    Get applications that failed with detailed error information.

    Args:
        profile: Profile name (optional)
    """
    try:
        from src.core.job_database import get_job_db
        import utils
        import sqlite3

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get job database
        db = get_job_db(profile_name)

        # Get failed applications with details
        with sqlite3.connect(db.db_path) as conn:
            conn.row_factory = sqlite3.Row

            cursor = conn.execute("""
                SELECT a.*, j.title, j.company, j.url, j.location
                FROM applications a
                JOIN jobs j ON a.job_id = j.id
                WHERE a.status = 'failed'
                ORDER BY a.applied_at DESC
            """)

            failed_apps = [dict(row) for row in cursor.fetchall()]

        # Parse JSON fields
        for app in failed_apps:
            for field in ['error_details', 'screening_questions', 'screening_answers']:
                if app.get(field):
                    try:
                        app[field] = json.loads(app[field])
                    except:
                        pass

        return {
            "failed_applications": failed_apps,
            "total": len(failed_apps),
            "profile": profile_name
        }

    except Exception as e:
        console.print(f"[red]Error getting failed applications: {e}[/red]")
        return {"error": str(e)}


@app.get("/api/realtime-errors")
async def get_realtime_errors(profile: str = None, severity: str = None):
    """
    Get real-time error tracking information.

    Args:
        profile: Profile name (optional)
        severity: Filter by severity level (optional)
    """
    try:
        from realtime_error_tracker import get_error_tracker
        import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get error tracker
        tracker = get_error_tracker(profile_name)

        # Get active errors
        active_errors = tracker.get_active_errors(severity)

        # Get error statistics
        statistics = tracker.get_error_statistics()

        return {
            "active_errors": active_errors,
            "statistics": statistics,
            "profile": profile_name,
            "monitoring_status": tracker.is_monitoring
        }

    except Exception as e:
        console.print(f"[red]Error getting realtime errors: {e}[/red]")
        return {"error": str(e)}


@app.post("/api/resolve-error/{error_id}")
async def resolve_error_endpoint(error_id: str, resolution_note: str = "", profile: str = None):
    """
    Resolve a real-time error.

    Args:
        error_id: Error ID to resolve
        resolution_note: Note about resolution
        profile: Profile name (optional)
    """
    try:
        from realtime_error_tracker import get_error_tracker
        import utils

        # Get available profiles
        profiles = utils.get_available_profiles()
        if not profiles:
            return {"error": "No profiles available"}

        # Use specified profile or first available
        profile_name = profile if profile in profiles else profiles[0]

        # Get error tracker
        tracker = get_error_tracker(profile_name)

        # Resolve the error
        success = tracker.resolve_error(error_id, resolution_note)

        if success:
            # Broadcast update via WebSocket
            await manager.broadcast({
                "type": "error_resolved",
                "error_id": error_id,
                "resolution_note": resolution_note,
                "timestamp": datetime.now().isoformat()
            })

            return {"success": True, "message": f"Error {error_id} resolved"}
        else:
            return {"error": f"Error {error_id} not found"}

    except Exception as e:
        console.print(f"[red]Error resolving error: {e}[/red]")
        return {"error": str(e)}


@app.get("/log")
async def log(limit: int = 100):
    """
    Get recent application logs.

    Args:
        limit: Maximum number of logs to return
    """
    return get_recent_logs(limit)


@app.post("/action/pause")
async def pause():
    """
    Pause the job application process.
    """
    success = set_pause_status(True)

    if success:
        # Broadcast pause status to WebSocket clients
        await manager.broadcast({"action": "pause", "status": True})
        return {"status": "paused"}
    else:
        raise HTTPException(status_code=500, detail="Failed to set pause status")

@app.post("/api/pause")
async def pause_automation():
    """
    Pause automation (alias for compatibility).
    """
    return await pause()


@app.post("/action/resume")
async def resume():
    """
    Resume the job application process.
    """
    success = set_pause_status(False)
    
    if success:
        # Broadcast resume status to WebSocket clients
        await manager.broadcast({"action": "pause", "status": False})
        return {"status": "resumed"}
    else:
        raise HTTPException(status_code=500, detail="Failed to set resume status")


@app.get("/status")
async def status():
    """
    Get the current pause status.
    """
    return {"paused": get_pause_status()}


# WebSocket endpoint
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    """
    WebSocket endpoint for real-time updates.
    """
    await manager.connect(websocket)
    
    try:
        # Send initial stats and status
        await websocket.send_json({
            "type": "stats",
            "data": get_application_stats()
        })
        
        await websocket.send_json({
            "type": "status",
            "paused": get_pause_status()
        })
        
        # Listen for messages
        while True:
            data = await websocket.receive_text()
            
            # Process commands from client
            try:
                message = json.loads(data)
                
                if message.get("command") == "get_stats":
                    await websocket.send_json({
                        "type": "stats",
                        "data": get_application_stats()
                    })
                
                elif message.get("command") == "get_logs":
                    limit = message.get("limit", 100)
                    await websocket.send_json({
                        "type": "logs",
                        "data": get_recent_logs(limit)
                    })
                
                elif message.get("command") == "get_status":
                    await websocket.send_json({
                        "type": "status",
                        "paused": get_pause_status()
                    })
                
            except json.JSONDecodeError:
                pass
            
    except WebSocketDisconnect:
        manager.disconnect(websocket)


# Event handlers
@app.on_event("startup")
async def startup_event():
    """
    Startup event handler.
    """
    console.print("[bold green]Dashboard API started[/bold green]")
    
    # Ensure IPC file exists
    if not IPC_FILE.exists():
        set_pause_status(False)


@app.on_event("shutdown")
async def shutdown_event():
    """
    Shutdown event handler.
    """
    console.print("[bold yellow]Dashboard API shutting down[/bold yellow]")


# Real-time Cache Endpoints
@app.get("/api/realtime-jobs")
async def get_realtime_jobs(
    limit: int = 50,
    profile: Optional[str] = None,
    keyword: Optional[str] = None,
    site: Optional[str] = None,
    hours: Optional[int] = 24
):
    """
    Get jobs from the real-time cache for instant dashboard updates.
    This endpoint provides near-instant access to the latest scraped jobs.
    """
    try:
        if not CACHE_AVAILABLE:
            return {
                "jobs": [],
                "cache_enabled": False,
                "message": "Real-time cache not available",
                "fallback_to_db": True
            }
        
        # Get jobs from cache based on filters
        if keyword:
            jobs = job_cache.get_jobs_by_keyword(keyword, limit)
        elif site:
            jobs = job_cache.get_jobs_by_site(site, limit)
        elif hours and hours > 0:
            jobs = job_cache.get_recent_jobs(hours)
            jobs = jobs[-limit:] if len(jobs) > limit else jobs
        else:
            jobs = job_cache.get_jobs(limit, profile)
        
        # Get cache stats
        cache_stats = get_cache_stats()
        
        return {
            "jobs": jobs,
            "cache_enabled": True,
            "cache_stats": cache_stats,
            "filters_applied": {
                "limit": limit,
                "profile": profile,
                "keyword": keyword,
                "site": site,
                "hours": hours
            },
            "total_returned": len(jobs)
        }
        
    except Exception as e:
        console.print(f"[red]Error getting real-time jobs: {e}[/red]")
        return {
            "jobs": [],
            "cache_enabled": False,
            "error": str(e),
            "fallback_to_db": True
        }

@app.get("/api/realtime-cache-stats")
async def get_realtime_cache_stats():
    """
    Get real-time cache statistics.
    """
    try:
        if not CACHE_AVAILABLE:
            return {
                "cache_enabled": False,
                "message": "Real-time cache not available"
            }
        
        stats = get_cache_stats()
        
        return {
            "cache_enabled": True,
            "stats": stats,
            "status": "active" if stats.get("cache_size", 0) > 0 else "empty"
        }
        
    except Exception as e:
        console.print(f"[red]Error getting cache stats: {e}[/red]")
        return {
            "cache_enabled": False,
            "error": str(e)
        }

@app.post("/api/realtime-cache/clear")
async def clear_realtime_cache():
    """
    Clear the real-time cache.
    """
    try:
        if not CACHE_AVAILABLE:
            return {
                "success": False,
                "message": "Real-time cache not available"
            }
        
        job_cache.clear()
        
        return {
            "success": True,
            "message": "Real-time cache cleared successfully",
            "cache_stats": get_cache_stats()
        }
        
    except Exception as e:
        console.print(f"[red]Error clearing cache: {e}[/red]")
        return {
            "success": False,
            "error": str(e)
        }

@app.get("/api/realtime-jobs/combined")
async def get_combined_jobs(
    limit: int = 50,
    profile: Optional[str] = None,
    use_cache: bool = True,
    cache_weight: float = 0.7
):
    """
    Get jobs combining real-time cache and database for optimal performance.
    This provides the best of both worlds - instant cache access and reliable database data.
    """
    try:
        combined_jobs = []
        cache_jobs = []
        db_jobs = []
        
        # Get jobs from cache if available and enabled
        if CACHE_AVAILABLE and use_cache:
            try:
                cache_jobs = job_cache.get_jobs(limit, profile)
                console.print(f"[blue]🚀 Retrieved {len(cache_jobs)} jobs from cache[/blue]")
            except Exception as e:
                console.print(f"[yellow]⚠️ Cache error: {e}[/yellow]")
        
        # Get jobs from database
        try:
            from src.core.job_database import get_job_db
            from src.core import utils
            
            # Get available profiles
            profiles = utils.get_available_profiles()
            
            # Use specified profile or first available
            profile_name = profile if profile else (profiles[0] if profiles else None)
            
            if profile_name:
                db = get_job_db(profile_name)
                db_jobs = db.get_jobs(limit=limit, offset=0)
                console.print(f"[green]💾 Retrieved {len(db_jobs)} jobs from database[/green]")
        except Exception as e:
            console.print(f"[yellow]⚠️ Database error: {e}[/yellow]")
        
        # Combine jobs with cache priority
        if cache_jobs and db_jobs:
            # Create sets for deduplication
            cache_urls = {job.get('url', '') for job in cache_jobs}
            
            # Add cache jobs first
            combined_jobs.extend(cache_jobs)
            
            # Add database jobs that aren't in cache
            for db_job in db_jobs:
                if db_job.get('url', '') not in cache_urls:
                    combined_jobs.append(db_job)
                    if len(combined_jobs) >= limit:
                        break
        elif cache_jobs:
            combined_jobs = cache_jobs
        elif db_jobs:
            combined_jobs = db_jobs
        
        # Limit the combined results
        combined_jobs = combined_jobs[:limit]
        
        return {
            "jobs": combined_jobs,
            "cache_enabled": CACHE_AVAILABLE,
            "cache_jobs_count": len(cache_jobs),
            "db_jobs_count": len(db_jobs),
            "combined_count": len(combined_jobs),
            "profile_used": profile_name,
            "cache_weight": cache_weight
        }
        
    except Exception as e:
        console.print(f"[red]Error getting combined jobs: {e}[/red]")
        return {
            "jobs": [],
            "error": str(e),
            "cache_enabled": CACHE_AVAILABLE
        }

# Job Filtering Endpoints
@app.get("/api/job-filters/stats")
async def get_job_filter_stats():
    """
    Get job filtering statistics and configuration.
    """
    try:
        if not FILTER_AVAILABLE:
            return {
                "filter_enabled": False,
                "message": "Job filtering not available"
            }
        
        stats = get_filter_stats()
        
        return {
            "filter_enabled": True,
            "stats": stats,
            "description": "French and Montreal job filtering system"
        }
        
    except Exception as e:
        console.print(f"[red]Error getting filter stats: {e}[/red]")
        return {
            "filter_enabled": False,
            "error": str(e)
        }

@app.post("/api/job-filters/test")
async def test_job_filter(job_data: Dict):
    """
    Test job filtering on a single job.
    """
    try:
        if not FILTER_AVAILABLE:
            return {
                "filter_enabled": False,
                "message": "Job filtering not available"
            }
        
        filter_result = filter_job(job_data)
        
        return {
            "filter_enabled": True,
            "job_title": job_data.get('title', 'Unknown'),
            "should_keep": filter_result.should_keep,
            "score": filter_result.score,
            "reasons": filter_result.reasons,
            "penalties": filter_result.penalties,
            "warnings": filter_result.warnings
        }
        
    except Exception as e:
        console.print(f"[red]Error testing job filter: {e}[/red]")
        return {
            "filter_enabled": False,
            "error": str(e)
        }

@app.post("/api/job-filters/batch")
async def filter_jobs_batch_endpoint(jobs: List[Dict]):
    """
    Filter a batch of jobs.
    """
    try:
        if not FILTER_AVAILABLE:
            return {
                "filter_enabled": False,
                "message": "Job filtering not available",
                "jobs": jobs  # Return original jobs if filtering not available
            }
        
        kept_jobs, filtered_jobs, stats = filter_jobs_batch(jobs)
        
        return {
            "filter_enabled": True,
            "kept_jobs": kept_jobs,
            "filtered_jobs": filtered_jobs,
            "stats": stats,
            "total_input": len(jobs),
            "kept_count": len(kept_jobs),
            "filtered_count": len(filtered_jobs)
        }
        
    except Exception as e:
        console.print(f"[red]Error filtering jobs batch: {e}[/red]")
        return {
            "filter_enabled": False,
            "error": str(e),
            "jobs": jobs  # Return original jobs if filtering fails
        }

@app.get("/api/jobs-filtered")
async def get_filtered_jobs(
    limit: int = 50,
    profile: Optional[str] = None,
    min_score: float = 50.0,
    show_filtered: bool = False
):
    """
    Get jobs with filtering applied.
    """
    try:
        # Get jobs from database
        from src.core.job_database import get_job_db
        from src.core import utils
        
        # Get available profiles
        profiles = utils.get_available_profiles()
        
        # Use specified profile or first available
        profile_name = profile if profile else (profiles[0] if profiles else None)
        
        if not profile_name:
            return {
                "jobs": [],
                "error": "No profile available",
                "filter_enabled": FILTER_AVAILABLE
            }
        
        db = get_job_db(profile_name)
        all_jobs = db.get_jobs(limit=limit*2, offset=0)  # Get more jobs to account for filtering
        
        if not FILTER_AVAILABLE:
            return {
                "jobs": all_jobs[:limit],
                "filter_enabled": False,
                "message": "Job filtering not available"
            }
        
        # Apply filtering
        kept_jobs, filtered_jobs, stats = filter_jobs_batch(all_jobs)
        
        # Filter by minimum score
        if min_score > 50.0:
            kept_jobs = [job for job in kept_jobs if job.get('filter_score', 100) >= min_score]
        
        # Limit results
        kept_jobs = kept_jobs[:limit]
        
        result = {
            "jobs": kept_jobs,
            "filter_enabled": True,
            "stats": stats,
            "min_score": min_score,
            "kept_count": len(kept_jobs),
            "filtered_count": len(filtered_jobs)
        }
        
        # Include filtered jobs if requested
        if show_filtered:
            result["filtered_jobs"] = filtered_jobs[:10]  # Show first 10 filtered jobs
        
        return result
        
    except Exception as e:
        console.print(f"[red]Error getting filtered jobs: {e}[/red]")
        return {
            "jobs": [],
            "error": str(e),
            "filter_enabled": FILTER_AVAILABLE
        }

# If this file is run directly, start the server
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8002)
