import json
import os
import shutil
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import openpyxl
from openpyxl.styles import Alignment, Font
from playwright.sync_api import Page
from rich.console import Console

# Initialize console for rich output
console = Console()

# Constants
PROFILES_DIR = Path("profiles")
OUTPUT_DIR = Path("output")
LOGS_DIR = OUTPUT_DIR / "logs"
IPC_FILE = OUTPUT_DIR / "ipc.json"

# Ensure directories exist
PROFILES_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
LOGS_DIR.mkdir(exist_ok=True)


class NeedsHumanException(Exception):
    """Exception raised when human intervention is required."""
    pass


def load_profile(profile_name: str) -> Dict:
    """
    Load a profile from the profiles directory.
    
    Args:
        profile_name: Name of the profile to load
        
    Returns:
        Profile dictionary
    """
    profile_dir = PROFILES_DIR / profile_name
    profile_file = profile_dir / f"{profile_name}.json"
    
    if not profile_dir.exists():
        raise FileNotFoundError(f"Profile directory not found: {profile_dir}")
    
    if not profile_file.exists():
        raise FileNotFoundError(f"Profile file not found: {profile_file}")
    
    try:
        with open(profile_file, "r") as f:
            profile = json.load(f)
        
        # Add profile name and directory to the profile
        profile["profile_name"] = profile_name
        profile["profile_dir"] = str(profile_dir)
        
        # Ensure required fields exist
        if "name" not in profile:
            profile["name"] = profile_name
        
        if "keywords" not in profile:
            profile["keywords"] = []
        
        if "batch_default" not in profile:
            profile["batch_default"] = 10
        
        if "ollama_model" not in profile:
            profile["ollama_model"] = "mistral:7b"
        
        # Set default document paths if not specified
        if "resume_docx" not in profile:
            profile["resume_docx"] = f"{profile_name}_Resume.docx"
        
        if "cover_letter_docx" not in profile:
            profile["cover_letter_docx"] = f"{profile_name}_CoverLetter.docx"
        
        if "resume_pdf" not in profile:
            profile["resume_pdf"] = f"{profile_name}_Resume.pdf"
        
        if "cover_letter_pdf" not in profile:
            profile["cover_letter_pdf"] = f"{profile_name}_CoverLetter.pdf"
        
        return profile
    
    except json.JSONDecodeError:
        raise ValueError(f"Invalid JSON in profile file: {profile_file}")


def ensure_profile_files(profile: Dict) -> None:
    """
    Ensure all required profile files exist.
    Generate PDFs if they don't exist or if DOCX is newer.
    
    Args:
        profile: Profile dictionary
    """
    profile_dir = Path(profile["profile_dir"])
    
    # Check for required DOCX files
    resume_docx = profile_dir / profile["resume_docx"]
    cover_letter_docx = profile_dir / profile["cover_letter_docx"]
    
    if not resume_docx.exists():
        raise FileNotFoundError(f"Resume DOCX not found: {resume_docx}")
    
    if not cover_letter_docx.exists():
        raise FileNotFoundError(f"Cover letter DOCX not found: {cover_letter_docx}")
    
    # Check for PDF files and generate if needed
    resume_pdf = profile_dir / profile["resume_pdf"]
    cover_letter_pdf = profile_dir / profile["cover_letter_pdf"]
    
    # Check if PDFs need to be generated (don't exist or DOCX is newer)
    resume_pdf_needed = (
        not resume_pdf.exists() or 
        resume_docx.stat().st_mtime > resume_pdf.stat().st_mtime
    ) if resume_pdf.exists() else True
    
    cover_letter_pdf_needed = (
        not cover_letter_pdf.exists() or 
        cover_letter_docx.stat().st_mtime > cover_letter_pdf.stat().st_mtime
    ) if cover_letter_pdf.exists() else True
    
    # Generate PDFs if needed
    if resume_pdf_needed:
        console.print(f"[yellow]Generating resume PDF...[/yellow]")
        success = convert_doc_to_pdf(str(resume_docx), str(resume_pdf))
        if not success:
            console.print(f"[bold red]Failed to generate resume PDF[/bold red]")
    
    if cover_letter_pdf_needed:
        console.print(f"[yellow]Generating cover letter PDF...[/yellow]")
        success = convert_doc_to_pdf(str(cover_letter_docx), str(cover_letter_pdf))
        if not success:
            console.print(f"[bold red]Failed to generate cover letter PDF[/bold red]")


def convert_doc_to_pdf(docx_path: str, pdf_path: str) -> bool:
    """
    Convert a DOCX file to PDF.
    Tries Word COM first, then falls back to other methods.
    
    Args:
        docx_path: Path to the DOCX file
        pdf_path: Path to save the PDF file
        
    Returns:
        True if conversion was successful, False otherwise
    """
    # Try Word COM first (Windows)
    try:
        success = convert_doc_to_pdf_windows(docx_path, pdf_path)
        if success:
            return True
    except Exception as e:
        console.print(f"[yellow]Word COM conversion failed: {e}[/yellow]")
    
    # TODO: Add fallback conversion methods here
    console.print("[yellow]No PDF conversion method available. Using DOCX instead.[/yellow]")
    
    # If all methods fail, copy the DOCX as a fallback
    try:
        shutil.copy2(docx_path, pdf_path)
        console.print(f"[yellow]Copied DOCX to {pdf_path} as fallback[/yellow]")
        return True
    except Exception as e:
        console.print(f"[bold red]Failed to copy DOCX as fallback: {e}[/bold red]")
        return False


def convert_doc_to_pdf_windows(docx_path: str, pdf_path: str) -> bool:
    """
    Convert a DOCX file to PDF using Word COM on Windows.
    
    Args:
        docx_path: Path to the DOCX file
        pdf_path: Path to save the PDF file
        
    Returns:
        True if conversion was successful, False otherwise
    """
    try:
        import comtypes.client
        import win32com.client
        
        # Constants for Word
        wdFormatPDF = 17
        
        # Create Word application
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False
        
        try:
            # Open the document
            doc = word.Documents.Open(os.path.abspath(docx_path))
            
            # Save as PDF
            doc.SaveAs(os.path.abspath(pdf_path), FileFormat=wdFormatPDF)
            
            # Close the document
            doc.Close()
            
            return True
            
        finally:
            # Quit Word application
            word.Quit()
    
    except ImportError:
        console.print("[yellow]Word COM libraries not available[/yellow]")
        return False
    except Exception as e:
        console.print(f"[bold red]Error converting to PDF: {e}[/bold red]")
        return False


def hash_job(job: Dict) -> str:
    """
    Create a hash for a job to identify duplicates.
    
    Args:
        job: Job dictionary
        
    Returns:
        Hash string for the job
    """
    # Create a string with key job attributes
    job_str = f"{job.get('title', '')}-{job.get('company', '')}-{job.get('location', '')}-{job.get('url', '')}"
    
    # Use a simple hash function
    import hashlib
    return hashlib.md5(job_str.encode()).hexdigest()[:10]


def load_session(profile: Dict, ats: str = None) -> Dict:
    """
    Load a session for a profile.
    
    Args:
        profile: Profile dictionary
        ats: ATS system to use (optional)
        
    Returns:
        Session dictionary
    """
    profile_dir = Path(profile["profile_dir"])
    session_file = profile_dir / "session.json"
    
    # Create default session if file doesn't exist
    if not session_file.exists():
        session = {
            "ats": ats if ats else "auto",
            "next_index": 0,
            "done": []
        }
        save_session(profile, session)
        return session
    
    # Load existing session
    try:
        with open(session_file, "r") as f:
            session = json.load(f)
        
        # Update ATS if provided
        if ats and ats != "resume":
            session["ats"] = ats
        
        return session
    
    except (json.JSONDecodeError, FileNotFoundError):
        # Create new session if file is invalid
        session = {
            "ats": ats if ats else "auto",
            "next_index": 0,
            "done": []
        }
        save_session(profile, session)
        return session


def save_session(profile: Dict, session: Dict) -> None:
    """
    Save a session for a profile.
    
    Args:
        profile: Profile dictionary
        session: Session dictionary
    """
    profile_dir = Path(profile["profile_dir"])
    session_file = profile_dir / "session.json"
    
    # Ensure profile directory exists
    profile_dir.mkdir(exist_ok=True)
    
    # Save session
    with open(session_file, "w") as f:
        json.dump(session, f, indent=2)


def check_pause_signal() -> bool:
    """
    Check if a pause signal has been set.
    
    Returns:
        True if pause signal is active, False otherwise
    """
    if not IPC_FILE.exists():
        return False
    
    try:
        with open(IPC_FILE, "r") as f:
            ipc_data = json.load(f)
        
        return ipc_data.get("pause", False)
    
    except (json.JSONDecodeError, FileNotFoundError):
        return False


def set_pause_signal(pause: bool) -> None:
    """
    Set the pause signal.
    
    Args:
        pause: True to pause, False to resume
    """
    with open(IPC_FILE, "w") as f:
        json.dump({"pause": pause}, f)


def append_log_row(
    job: Dict,
    profile: Dict,
    status: str,
    resume_path: str,
    cover_letter_path: str,
    ats: str
) -> None:
    """
    Append a row to the applications log.
    
    Args:
        job: Job dictionary
        profile: Profile dictionary
        status: Application status
        resume_path: Path to the resume file
        cover_letter_path: Path to the cover letter file
        ats: ATS system used
    """
    log_file = LOGS_DIR / "applications.xlsx"
    
    # Create log file if it doesn't exist
    if not log_file.exists():
        create_log_file(log_file)
    
    # Try to open the log file (with retries for file locks)
    for attempt in range(5):
        try:
            wb = openpyxl.load_workbook(log_file)
            ws = wb.active
            
            # Get the next row
            next_row = ws.max_row + 1
            
            # Add the log entry
            ws.cell(row=next_row, column=1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=next_row, column=2).value = profile.get("name", "")
            ws.cell(row=next_row, column=3).value = job.get("title", "")
            ws.cell(row=next_row, column=4).value = job.get("company", "")
            ws.cell(row=next_row, column=5).value = job.get("location", "")
            ws.cell(row=next_row, column=6).value = job.get("url", "")
            ws.cell(row=next_row, column=7).value = status
            ws.cell(row=next_row, column=8).value = os.path.basename(resume_path) if resume_path else ""
            ws.cell(row=next_row, column=9).value = os.path.basename(cover_letter_path) if cover_letter_path else ""
            ws.cell(row=next_row, column=10).value = ats
            ws.cell(row=next_row, column=11).value = hash_job(job)
            
            # Save the workbook
            wb.save(log_file)
            break
            
        except PermissionError:
            if attempt < 4:
                console.print(f"[yellow]Log file is locked. Retrying in {attempt + 1} seconds...[/yellow]")
                time.sleep(attempt + 1)
            else:
                # If all retries fail, write to fallback CSV
                write_fallback_log(job, profile, status, resume_path, cover_letter_path, ats)
        
        except Exception as e:
            console.print(f"[bold red]Error writing to log: {e}[/bold red]")
            write_fallback_log(job, profile, status, resume_path, cover_letter_path, ats)
            break


def create_log_file(log_file: Path) -> None:
    """
    Create a new applications log file.
    
    Args:
        log_file: Path to the log file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    
    # Add headers
    headers = [
        "Timestamp", "Profile", "Job Title", "Company", "Location", "URL",
        "Status", "Resume", "Cover Letter", "ATS", "Hash"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Save the workbook
    wb.save(log_file)


def write_fallback_log(
    job: Dict,
    profile: Dict,
    status: str,
    resume_path: str,
    cover_letter_path: str,
    ats: str
) -> None:
    """
    Write to a fallback CSV log when Excel fails.
    
    Args:
        job: Job dictionary
        profile: Profile dictionary
        status: Application status
        resume_path: Path to the resume file
        cover_letter_path: Path to the cover letter file
        ats: ATS system used
    """
    fallback_log = LOGS_DIR / "fallback.csv"
    
    # Create headers if file doesn't exist
    if not fallback_log.exists():
        with open(fallback_log, "w") as f:
            f.write("Timestamp,Profile,Job Title,Company,Location,URL,Status,Resume,Cover Letter,ATS,Hash\n")
    
    # Escape commas in fields
    def escape(s):
        if isinstance(s, str):
            return '"' + s.replace('"', '""') + '"'
        return str(s)
    
    # Prepare log entry
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    profile_name = profile.get("name", "")
    title = job.get("title", "")
    company = job.get("company", "")
    location = job.get("location", "")
    url = job.get("url", "")
    resume = os.path.basename(resume_path) if resume_path else ""
    cover_letter = os.path.basename(cover_letter_path) if cover_letter_path else ""
    job_hash = hash_job(job)
    
    # Write log entry
    with open(fallback_log, "a") as f:
        f.write(f"{escape(timestamp)},{escape(profile_name)},{escape(title)},{escape(company)},"
                f"{escape(location)},{escape(url)},{escape(status)},{escape(resume)},"
                f"{escape(cover_letter)},{escape(ats)},{escape(job_hash)}\n")


def save_customized_document(profile: Dict, job_hash: str, docx_path: str, is_resume: bool) -> str:
    """
    Save a customized document to the output directory.
    
    Args:
        profile: Profile dictionary
        job_hash: Hash of the job
        docx_path: Path to the DOCX file
        is_resume: True if this is a resume, False for cover letter
        
    Returns:
        Path to the saved document
    """
    profile_name = profile.get("profile_name", "unknown")
    doc_type = "Resume" if is_resume else "CoverLetter"
    
    # Create output directory
    output_dir = OUTPUT_DIR / profile_name / "documents"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Create output path
    output_path = output_dir / f"{job_hash}_{doc_type}.pdf"
    
    # Convert to PDF
    success = convert_doc_to_pdf(docx_path, str(output_path))
    
    # Return the path (either PDF or original DOCX if conversion failed)
    return str(output_path if success else docx_path)


def smart_attach(page: Page, selectors: List[str], file_path: str) -> bool:
    """
    Try multiple selectors to attach a file.
    
    Args:
        page: Playwright page
        selectors: List of selectors to try
        file_path: Path to the file to attach
        
    Returns:
        True if successful, False otherwise
        
    Raises:
        NeedsHumanException: If no selector matches
    """
    for selector in selectors:
        try:
            # Try to find the element
            element = page.query_selector(selector)
            if element:
                # Attach the file
                element.set_input_files(file_path)
                return True
        except Exception:
            continue
    
    # If no selector worked, raise exception
    raise NeedsHumanException(f"Could not find file input. Please attach manually: {file_path}")


def fill_if_empty(page: Page, selector: str, value: str) -> bool:
    """
    Fill a field only if it's empty.
    
    Args:
        page: Playwright page
        selector: Selector for the field
        value: Value to fill
        
    Returns:
        True if field was filled, False otherwise
    """
    try:
        element = page.query_selector(selector)
        if element:
            # Check if the field is empty
            current_value = element.get_attribute("value") or ""
            if not current_value.strip():
                # Fill the field
                element.fill(value)
                return True
    except Exception:
        pass
    
    return False


def loop_click(page: Page, selector: str, max_clicks: int = 5) -> int:
    """
    Click a button repeatedly until it's no longer visible.
    
    Args:
        page: Playwright page
        selector: Selector for the button
        max_clicks: Maximum number of clicks
        
    Returns:
        Number of clicks performed
    """
    clicks = 0
    
    for _ in range(max_clicks):
        try:
            # Check if the button is visible
            if page.is_visible(selector):
                # Click the button
                page.click(selector)
                clicks += 1
                
                # Wait for navigation or processing
                page.wait_for_timeout(1000)
            else:
                break
        except Exception:
            break
    
    return clicks


def create_temp_file(prefix: str = "autojob_", suffix: str = ".docx") -> str:
    """
    Create a temporary file path.
    
    Args:
        prefix: Prefix for the filename
        suffix: Suffix for the filename
        
    Returns:
        Path to the temporary file
    """
    temp_dir = Path("temp")
    temp_dir.mkdir(exist_ok=True)
    
    # Create a unique filename
    filename = f"{prefix}{uuid.uuid4().hex}{suffix}"
    
    return str(temp_dir / filename)
